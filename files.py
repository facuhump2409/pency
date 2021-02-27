import datetime
import os

from openpyxl import *

from model.commons import load_products_in_upper_case

excels = []
dir_path = os.path.dirname(os.path.realpath(__file__))
now = datetime.datetime.now()
month, year = str(now.month), str(now.year)
month = "0" + month if len(month) == 1 else month
# TODO agregar el mes a las carpetas y archivos
orders_folder_year = dir_path + "/Pedidos/" + year + "/012020/"
clients_folder = orders_folder_year + "Generador de pedidos/Pedidos/"
orders_generator = load_workbook(orders_folder_year + "Generador de pedidos/Generador de pedidos.xlsx")
consolidated_orders_with_formulas = load_workbook(orders_folder_year + "Pedidos consolidados 01-" + year + ".xlsx")
consolidated_orders_data_only = load_workbook(orders_folder_year + "Pedidos consolidados 01-" + year + ".xlsx",
                                              data_only=True)
prices = consolidated_orders_data_only['Precios y Menú']
excels.extend([consolidated_orders_with_formulas, consolidated_orders_data_only, orders_generator])

products_dictionary = {}  # para saber cual es premium, guarnicion o daily
load_products_in_upper_case(products_dictionary, prices)

orders_messages = orders_generator["Nuevos mensajes"]
