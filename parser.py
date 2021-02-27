import re
import os
import datetime

from openpyxl import *
from model.commons import load_products, add_atributes_to_excels
from model.pedido import Plato, get_basic_information, work_products

excels = []
dir_path = os.path.dirname(os.path.realpath(__file__))
now = datetime.datetime.now()
month,year = str(now.month), str(now.year)
month = "0" + month if len(month) == 1 else month
# TODO agregar el mes a las carpetas y archivos
orders_folder_year = dir_path + "/Pedidos/" + year + "/012020/"
clients_folder = orders_folder_year + "Generador de pedidos/Pedidos/"
orders_generator = load_workbook(orders_folder_year + "Generador de pedidos/Generador de pedidos.xlsx")
consolidated_orders_with_formulas = load_workbook(orders_folder_year + "Pedidos consolidados 01-" + year + ".xlsx")
consolidated_orders_data_only = load_workbook(orders_folder_year + "Pedidos consolidados 01-" + year + ".xlsx", data_only=True)
prices = consolidated_orders_data_only['Precios y Menú']
excels.extend([consolidated_orders_with_formulas, consolidated_orders_data_only, orders_generator])
products_dictionary = {}  # para saber cual es premium, guarnicion o daily

load_products(products_dictionary, prices)

messages = orders_generator["Mensajes de wapp"]


def process_orders(messages):
    message_row = 0
    for message in messages.iter_rows(values_only=True):
        try:
            message_row += 1
            if message[0] == 'Si' or message[0] == 'Procesado': continue
            client_order_data_only = load_workbook(clients_folder + "/Planilla Cliente.xlsx", data_only=True)
            client_order_with_formula = load_workbook(clients_folder + "/Planilla Cliente.xlsx")
            excel_dictionary = {}
            order = message[1]
            get_basic_information(excel_dictionary, order)
            items_regex = "(?i)— .+"
            items = re.findall(items_regex, order)
            order_products = []
            guarnicion_regex = "(?i)Guarnici.n(es)?"
            for item in items:
                qty_regex = "\[ \d \]"
                qty_product = re.search(qty_regex, item)
                qty = 1 if qty_product == None else int(re.sub("\[|\]", "", qty_product[0]).strip())
                product = re.sub(r"" + qty_regex + "|" + guarnicion_regex + ".+", "",
                                 re.search("(?i)(?<=—).+(?=\>)", item)[0].strip())
                extra_brackets_info = "\s?(\(.+)?:"
                cleaner_item = re.search(guarnicion_regex + extra_brackets_info + " .+(?=>)", item)[0]
                guarniciones = re.sub(guarnicion_regex + extra_brackets_info, "", cleaner_item).strip()
                guarniciones = re.findall('[A-W][^A-W]*', guarniciones)
                guarniciones_list = []
                for guarnicion in guarniciones:
                    nombre = re.sub(guarnicion_regex + "?: |X\d|,", "", guarnicion).strip()
                    guarnicion_qty = re.search("X\d", guarnicion)
                    final_qty = 1 if guarnicion_qty == None else int(re.sub("X", "", guarnicion_qty[0]).strip())
                    guarniciones_list.append(Plato(final_qty, nombre))
                work_products(product, order_products, qty, guarniciones_list, products_dictionary)
            # excel_dictionary['Zona'] = re.search("(?i)(?<=Zona de entrega: ).+",order)[0]
            messages.cell(message_row, 1).value = "Si"
            add_atributes_to_excels([[client_order_data_only, client_order_with_formula],
                                     [consolidated_orders_data_only, consolidated_orders_with_formulas]],
                                    excel_dictionary,
                                    order_products)
            client_order_with_formula.save(
                clients_folder + excel_dictionary['Pedido'] + "_" + excel_dictionary['Cliente'] + ".xlsx")
            client_order_with_formula.close()
            client_order_data_only.close()
        except:
            messages.cell(message_row, 1).value = "ERROR"
    orders_generator.save(orders_folder_year + "/Generador de pedidos/Generador de pedidos.xlsx")
    consolidated_orders_with_formulas.save(orders_folder_year + "Pedidos consolidados 01-" + year + ".xlsx")
    for excel in excels:
        excel.close()


process_orders(messages)
# TODO ver convensiones de nombres de hojas de excel
